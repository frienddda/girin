from django.shortcuts import render, get_object_or_404, redirect
##from .models import jumun_T
##from .models import Document
#mysql
from .models import JumunT
from .models import OrderT
#모델 필드변경
##from .forms import RegisterForm
from django.utils import timezone
import xlwt
from django.http import HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q
import openpyxl
from django.core.files.storage import FileSystemStorage
from django.views.decorators.csrf import csrf_exempt
import simplejson as json
from django.contrib import messages

from django.contrib.auth.decorators import login_required

##from django.contrib.auth.decorators import permission_required



# Create your views here.


##def index(request):
##    return HttpResponse("jumun data")


@login_required(login_url='common:login')
def index(request):
    #페이지
    page = request.GET.get('page','1')

    #검색어
    kw = request.GET.get('kw', '')
    mkw1 = request.GET.get('mkw1', '').strip()
    mkw2 = request.GET.get('mkw2', '').strip()
    mkw3 = request.GET.get('mkw3', '').strip()
    mkw4 = request.GET.get('mkw4', '').strip()
    mkw5 = request.GET.get('mkw5', '').strip()
    mkc1 = request.GET.get('mkc1', '').strip()
    mkc2 = request.GET.get('mkc2', '').strip()
    mkc3 = request.GET.get('mkc3', '').strip()
    mkc4 = request.GET.get('mkc4', '').strip()
    mkc5 = request.GET.get('mkc5', '').strip()

    #계정정보저장
    username= request.user.username
    
    
    #조회
    if username == 'admin':
        jumun_list = JumunT.objects.order_by('jumun_t_id')
    else: #계정정보=주문자id 인 내용만 조회
        jumun_list =JumunT.objects.filter(Q(주문자id__icontains= username))
    
    all_c = jumun_list.count()

    if kw:
        jumun_list = jumun_list.filter(
            Q(주문자id__icontains=kw) | # 주문id
            Q(주문자__icontains=kw) |  # 주문자명
            Q(위탁자__icontains=kw) |  # 위탁
            Q(브랜드__icontains=kw)  # 브랜드
        )

    elif mkw1 or mkw2 or mkw3 or mkw4 or mkw5:
        if mkc1 or mkc2 or mkc3 or mkc4 or mkc5:
            jumun_list = jumun_list.filter(
                Q(주문자__in=[mkw1,mkw2,mkw3,mkw4,mkw5]),  # 주문자
                Q(위탁자__in=[mkc1,mkc2,mkc3,mkc4,mkc5])   # 위탁자
            
            )
        else:
            jumun_list = jumun_list.filter(
                Q(주문자__in=[mkw1,mkw2,mkw3,mkw4,mkw5]))
                
    kw_c = jumun_list.count()
    #jumun_list2 = 전체 조회내용
    #페이징
    paginator = Paginator(jumun_list, 20) #페이지당 20개
    page_obj = paginator.get_page(page)
    context = {'jumun_list':page_obj, 'jumun_list2':jumun_list, 'page': page, 'kw': kw, 'all_c':all_c, 'kw_c':kw_c,'mkw1':mkw1,'mkw2':mkw2,'mkw3':mkw3,'mkw4':mkw4,'mkw5':mkw5,'mkc1':mkc1,'mkc2':mkc2,'mkc3':mkc3,'mkc4':mkc4,'mkc5':mkc5}
##    context = {'jumun_list': jumun_list}
    return render(request, 'jumun/jumun_list.html', context)

    
    


#미사용 수정안했음

def jumun_create(request):

 
    if request.method == 'POST':
       form = RegisterForm(request.POST)
       if form.is_valid():
           jumun_T = form.save(commit=False)
           jumun_T.jumun_date = timezone.now()
           jumun_T.save()
           return redirect('jumun:index')
    else:
        form = RegisterForm()
    context = {'form' : form}
    return render(request, 'jumun/jumun_form.html', {'form': form})

##@permission_required('jumun.delete_jumun_t', raise_exception = True)
def jumun_delete(request):
    
    id_list2 = request.POST.getlist('selected[]')
    page = request.GET.get('page')    
    for d_id in id_list2:
        d = JumunT.objects.get(jumun_t_id=d_id)
        d.delete()
    
    return redirect('jumun:index')


## 테이블수정 "이름" 컬럼 test중 6월 23
@csrf_exempt
def jumun_td_update(request):

    data =json.loads(request.body)
    t_gubun =data.get('gubun', None)
    if t_gubun == "e_brand" :
        t_id = data.get('t_id', None)
        t_brand = data.get('t_brand', None)
        JumunT.objects.filter(jumun_t_id=t_id).update(브랜드=t_brand)
    elif t_gubun == "e_quantity":
        t_id = data.get('t_id', None)
        t_quantity = data.get('t_quantity', None)
        JumunT.objects.filter(jumun_t_id=t_id).update(수량=t_quantity)
    elif t_gubun == "e_memo":
        t_id = data.get('t_id', None)
        t_memo = data.get('t_memo', None)
        JumunT.objects.filter(jumun_t_id=t_id).update(비고=t_memo)
    elif t_gubun == "e_phone":
        t_id = data.get('t_id', None)
        t_phone = data.get('t_phone', None)
        JumunT.objects.filter(jumun_t_id=t_id).update(전화번호=t_phone)
        
    return redirect('jumun:index')
    ##return Response()

## 페이징안된 전체 검색리스트 다운로드
def jumun_excel2(request):
	
    response = HttpResponse(content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = 'attachment;filename*=UTF-8\'\'all_download.xls'


    wb = xlwt.Workbook(encoding='ansi') #encoding은 ansi로 해준다.
    ws = wb.add_sheet('주문장') #시트 추가
    
    row_num = 0
    col_names = ['영문', '주문일','주문ID','주문자','위탁자','브랜드','상품명','색상','수량','중도매','도매가','비고','공급가','이름','전화번호','주소','ID','나온날짜']
    
    #열이름을 첫번째 행에 추가 시켜준다.
    for idx, col_name in enumerate(col_names):
        ws.write(row_num, idx, col_name)


    id_list2 = request.POST.getlist('jumunlist[]')

    
    
    for d_id in id_list2:
        rows = JumunT.objects.filter(jumun_t_id=d_id).values_list('영문', '주문날짜','주문자id','주문자','위탁자','브랜드','상품명','색상','수량','중도매','도매가','비고','공급가','이름','전화번호','주소','아이디','나온날짜')
        

        for row in rows:
            row_num +=1
            for col_num, attr in enumerate(row):
                ws.write(row_num, col_num, attr)
    wb.save(response)
    return response
    

##선택 다운로드
def jumun_excel(request):
	
    response = HttpResponse(content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = 'attachment;filename*=UTF-8\'\'select_download.xls'

##    id_list = request.POST.getlist('selected[]')
    wb = xlwt.Workbook(encoding='ansi') #encoding은 ansi로 해준다.
    ws = wb.add_sheet('주문장') #시트 추가
    
    row_num = 0
    col_names = ['영문', '주문일','주문ID','주문자','위탁자','브랜드','상품명','색상','수량','중도매','도매가','비고','공급가','이름','전화번호','주소','ID','나온날짜']
    
    #열이름을 첫번째 행에 추가 시켜준다.
    for idx, col_name in enumerate(col_names):
        ws.write(row_num, idx, col_name)


    id_list2 = request.POST.getlist('selected[]')


    
    
    for d_id in id_list2:
        rows = JumunT.objects.filter(jumun_t_id=d_id).values_list('영문', '주문날짜','주문자id','주문자','위탁자','브랜드','상품명','색상','수량','중도매','도매가','비고','공급가','이름','전화번호','주소','아이디','나온날짜')
        

        for row in rows:
            row_num +=1
            for col_num, attr in enumerate(row):
                ws.write(row_num, col_num, attr)
    wb.save(response)
    return response


# 파일업로드 구현부
##def uploadFile(request):
##    if request.method == "POST":
##        # Fetching the form data
##        fileTitle = request.POST["fileTitle"]
##        uploadedFile = request.FILES["uploadedFile"]
##
##        # Saving the information in the database
##        document = Document(
##            title=fileTitle,
##            uploadedFile=uploadedFile
##        )
##        document.save()
##
##    documents = Document.objects.all()
##
####    return render(request, 'jumun/jumun_list.html', context={"files": documents})
##    return redirect('jumun:index')


    
##@csrf_exempt
#엑셀업로드 및 db저장
def excel_upload(request):
    if request.method == 'POST':

        # 파일 저장
        file = request.FILES['file_excel']
        fs = FileSystemStorage()
        filename = fs.save(file.name, file)
        uploaded_file_url = fs.url(filename)
        # print(uploaded_file_url)

        excel = openpyxl.load_workbook(file, data_only=True)

        work_sheet = excel.worksheets[0]

        #엑셀파일 마지막로우 체크
        cc= work_sheet.max_row
        cc = str(cc)

        rows = work_sheet['A2':'R'+cc]

        for row in rows:
            
            dict = {}
            dict['영문']=row[0].value
            dict['주문날짜']=row[1].value
            dict['주문자id']=row[2].value
            dict['주문자']=row[3].value
            dict['위탁자']=row[4].value
            dict['브랜드']=row[5].value
            dict['상품명']=row[6].value
            dict['색상']=row[7].value
            dict['수량']=row[8].value
            dict['중도매']=row[9].value
            dict['도매가']=row[10].value
            dict['비고']=row[11].value
            dict['공급가']=row[12].value
            dict['이름']=row[13].value
            dict['전화번호']=row[14].value
            dict['주소']=row[15].value
            dict['아이디']=row[16].value
            dict['나온날짜']=row[17].value

            JumunT(**dict).save()

        context = {'status': True, 'rtnmsg': '엑셀파일이 정상적으로 업로드 됐습니다.'}
        return HttpResponse(json.dumps(context), content_type='application/json')



def order_submit(request):
    id_list = request.POST.getlist('selected[]')
    

    for d_id in id_list:
        rows = JumunT.objects.filter(jumun_t_id=d_id).values('영문', '주문날짜','주문자id','주문자','위탁자','브랜드','상품명','색상','수량','중도매','도매가','비고','공급가','이름','전화번호','주소','아이디','나온날짜')
        
        for row in rows:
               
            dict = {}
            
            dict['영문']=row['영문']
            dict['주문날짜']=row['주문날짜']
            dict['주문자id']=row['주문자id']
            dict['주문자']=row['주문자']
            dict['위탁자']=row['위탁자']
            dict['브랜드']=row['브랜드']
            dict['상품명']=row['상품명']
            dict['색상']=row['색상']
            dict['수량']=row['수량']
            dict['중도매']=row['중도매']
            dict['도매가']=row['도매가']
            dict['비고']=row['비고']
            dict['공급가']=row['공급가']
            dict['이름']=row['이름']
            dict['전화번호']=row['전화번호']
            dict['주소']=row['주소']
            dict['아이디']=row['아이디']
            dict['나온날짜']=row['나온날짜']

            OrderT(**dict).save()
       #발주된 데이터 주문테이블에서 삭제
        d = JumunT.objects.get(jumun_t_id=d_id)
        d.delete()
    
    
    return redirect('jumun:index')

        

def all_delete(request):
    jumun_list = JumunT.objects.all()
    jumun_list.delete()
    messages.info(request, '모든 데이터가 삭제되었습니다.')
    return redirect('jumun:index')

def order_list(request):
    #페이지
    page = request.GET.get('page','1')

    #검색어
    kw = request.GET.get('kw', '').strip()
    
    mkw1 = request.GET.get('mkw1', '').strip()
    mkw2 = request.GET.get('mkw2', '').strip()
    mkw3 = request.GET.get('mkw3', '').strip()
    mkw4 = request.GET.get('mkw4', '').strip()
    mkw5 = request.GET.get('mkw5', '').strip()
    mkc1 = request.GET.get('mkc1', '').strip()
    mkc2 = request.GET.get('mkc2', '').strip()
    mkc3 = request.GET.get('mkc3', '').strip()
    mkc4 = request.GET.get('mkc4', '').strip()
    mkc5 = request.GET.get('mkc5', '').strip()

    #계정정보저장
    username= request.user.username
    
    
    #조회
    if username == 'admin':
        order_list = OrderT.objects.order_by('order_t_id')
    else: #계정정보=주문자id 인 내용만 조회
        order_list =OrderT.objects.filter(Q(주문자id__icontains= username))
    
    all_c = order_list.count()

    if kw:
        order_list = order_list.filter(
            Q(주문자id__icontains=kw) | # 주문id
            Q(주문자__icontains=kw) |  # 주문자명
            Q(위탁자__icontains=kw) |  # 위탁
            Q(브랜드__icontains=kw)  # 브랜드
        )

    elif mkw1 or mkw2 or mkw3 or mkw4 or mkw5:
        if mkc1 or mkc2 or mkc3 or mkc4 or mkc5:
            order_list = order_list.filter(
                Q(주문자__in=[mkw1,mkw2,mkw3,mkw4,mkw5]),  # 주문자
                Q(위탁자__in=[mkc1,mkc2,mkc3,mkc4,mkc5])   # 위탁자
            
            )
        else:
            order_list = order_list.filter(
                Q(주문자__in=[mkw1,mkw2,mkw3,mkw4,mkw5]))
                
    kw_c = order_list.count()
    #jumun_list2 = 전체 조회내용
    #페이징
    paginator = Paginator(order_list, 20) #페이지당 20개
    page_obj = paginator.get_page(page)
    context = {'order_list':page_obj, 'order_list2':order_list, 'page': page, 'kw': kw, 'all_c':all_c, 'kw_c':kw_c,'mkw1':mkw1,'mkw2':mkw2,'mkw3':mkw3,'mkw4':mkw4,'mkw5':mkw5,'mkc1':mkc1,'mkc2':mkc2,'mkc3':mkc3,'mkc4':mkc4,'mkc5':mkc5}

    return render(request, 'jumun/order_list.html', context)

def order_delete(request):
    
    id_list2 = request.POST.getlist('selected[]')
    page = request.GET.get('page')    
    for d_id in id_list2:
        d = OrderT.objects.get(order_t_id=d_id)
        d.delete()
    
    return redirect('jumun:order')



##선택 다운로드
def order_excel(request):
	
    response = HttpResponse(content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = 'attachment;filename*=UTF-8\'\'select_download.xls'


    wb = xlwt.Workbook(encoding='ansi') #encoding은 ansi로 해준다.
    ws = wb.add_sheet('order') #시트 추가
    
    row_num = 0
    col_names = ['영문', '주문일','주문ID','주문자','위탁자','브랜드','상품명','색상','수량','중도매','도매가','비고','공급가','이름','전화번호','주소','ID','나온날짜']
    
    #열이름을 첫번째 행에 추가 시켜준다.
    for idx, col_name in enumerate(col_names):
        ws.write(row_num, idx, col_name)


    id_list2 = request.POST.getlist('selected[]')


    
    
    for d_id in id_list2:
        rows = OrderT.objects.filter(order_t_id=d_id).values_list('영문', '주문날짜','주문자id','주문자','위탁자','브랜드','상품명','색상','수량','중도매','도매가','비고','공급가','이름','전화번호','주소','아이디','나온날짜')
        

        for row in rows:
            row_num +=1
            for col_num, attr in enumerate(row):
                ws.write(row_num, col_num, attr)
    wb.save(response)
    return response



## 페이징안된 전체 검색리스트 다운로드
def order_excel2(request):
	
    response = HttpResponse(content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = 'attachment;filename*=UTF-8\'\'all_download.xls'


    wb = xlwt.Workbook(encoding='ansi') #encoding은 ansi로 해준다.
    ws = wb.add_sheet('주문장') #시트 추가
    
    row_num = 0
    col_names = ['영문', '주문일','주문ID','주문자','위탁자','브랜드','상품명','색상','수량','중도매','도매가','비고','공급가','이름','전화번호','주소','ID','나온날짜']
    
    #열이름을 첫번째 행에 추가 시켜준다.
    for idx, col_name in enumerate(col_names):
        ws.write(row_num, idx, col_name)


    id_list2 = request.POST.getlist('orderlist[]')

    
    
    for d_id in id_list2:
        rows = OrderT.objects.filter(order_t_id=d_id).values_list('영문', '주문날짜','주문자id','주문자','위탁자','브랜드','상품명','색상','수량','중도매','도매가','비고','공급가','이름','전화번호','주소','아이디','나온날짜')
        

        for row in rows:
            row_num +=1
            for col_num, attr in enumerate(row):
                ws.write(row_num, col_num, attr)
    wb.save(response)
    return response


def order_all_delete(request):
    order_list = OrderT.objects.all()
    order_list.delete()
    messages.info(request, '모든 데이터가 삭제되었습니다.')
    return redirect('jumun:order')


    
    
